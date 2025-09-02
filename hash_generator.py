# hash_generator.py — sorties .xlsm (macro-enabled)
# -*- coding: utf-8 -*-
import os
import csv
import hashlib
import openpyxl
from openpyxl.styles import Protection
from integrity import stamp_workbook  # stamp_workbook(wb, template_version, student_id, main_sheet_name)

DATA_DIR = os.environ.get("DATA_DIR", "./")

def _safe_filename(s: str) -> str:
    keep = []
    for ch in s or "":
        if ch.isalnum() or ch in ("-", "_"):
            keep.append(ch)
        elif ch == " ":
            keep.append("_")
    return "".join(keep) or "file"

def _hash_contenu(ws, student_id: str) -> str:
    contenu = (student_id or "").encode()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                contenu += str(cell).encode()
    return hashlib.sha256(contenu).hexdigest()

def generate_student_files_csv(
    input_csv="liste_etudiants.csv",
    template_path="Fichier_Excel_Professeur_Template.xlsm",  # <<< .xlsm par défaut
    output_folder="copies_generees",                         # même dossier que côté prof
    log_file="hash_records.csv",
    template_version="v1.0.0",
):
    # Résolution via DATA_DIR
    input_csv     = input_csv     if os.path.isabs(input_csv)     else os.path.join(DATA_DIR, input_csv)
    template_path = template_path if os.path.isabs(template_path) else os.path.join(DATA_DIR, template_path)
    output_folder = output_folder if os.path.isabs(output_folder) else os.path.join(DATA_DIR, output_folder)
    log_file      = log_file      if os.path.isabs(log_file)      else os.path.join(DATA_DIR, log_file)

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template introuvable: {template_path}")
    os.makedirs(output_folder, exist_ok=True)

    # Charger la liste d'étudiants
    etudiants = []
    with open(input_csv, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            sid = (row.get("id") or "").strip()
            if not sid:
                continue
            etudiants.append({
                "id": sid,
                "nom": (row.get("nom") or "").strip(),
                "prenom": (row.get("prenom") or "").strip(),
            })

    # Log CSV
    with open(log_file, "w", newline="", encoding="utf-8") as flog:
        w = csv.writer(flog)
        w.writerow(["id_etudiant", "nom", "prenom", "hash", "nom_fichier"])

        for etu in etudiants:
            uid, nom, prenom = etu["id"], etu["nom"], etu["prenom"]

            # Charger le modèle en conservant les macros
            wb = openpyxl.load_workbook(template_path, keep_vba=True, data_only=False)
            ws = wb.active
            main_sheet_name = ws.title

            # Z1 / Z2
            ws["Z1"] = uid
            h = _hash_contenu(ws, uid)
            ws["Z2"] = h

            # Masquer Z
            ws.column_dimensions["Z"].hidden = True

            # Déverrouiller C..Y et protéger
            for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row or 2, min_col=3, max_col=25):
                for cell in row_cells:
                    cell.protection = Protection(locked=False)
            ws.protection.sheet = True
            ws.protection.enable()
            ws.protection.selectLockedCells = False
            ws.protection.selectUnlockedCells = True

            # Estampille d’intégrité
            stamp_workbook(wb, template_version=template_version, student_id=uid, main_sheet_name=main_sheet_name)

            # >>> Sauvegarde en .xlsm
            fname = f"{uid}_{_safe_filename(nom)}_{_safe_filename(prenom)}.xlsm"
            out_path = os.path.join(output_folder, fname)
            wb.save(out_path)

            w.writerow([uid, nom, prenom, h, fname])
            print(f"✅ Copie générée : {fname}")

    return os.path.abspath(output_folder)

if __name__ == "__main__":
    generate_student_files_csv(
        input_csv="liste_etudiants.csv",
        template_path="Fichier_Excel_Professeur_Template.xlsm",
        output_folder="copies_generees",
        log_file="hash_records.csv",
        template_version="v1.0.0",
    )

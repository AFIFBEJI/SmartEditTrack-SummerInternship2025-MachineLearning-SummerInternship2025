# integrity.py — estampillage d’intégrité (template + cellules)
import os, hmac, json, hashlib
from typing import Dict, List, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime

# ------------- Config -------------
SECRET = os.environ.get("SET_SECRET", "change-me-please")  # mets un secret fort en prod
SIG_SHEET = "_sig"                                         # feuille VeryHidden
RANGE_COLS = ("C", "Y")                                    # plage à signer
START_ROW = 2                                              # premières réponses
HEADER_KEY = "__header__"                                  # clé spéciale pour l'en-tête

# ------------- Utils -------------
def _h(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest()

def _hmac(payload: str) -> str:
    return hmac.new(SECRET.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()

def _cell_range(ws: Worksheet) -> List[str]:
    from openpyxl.utils.cell import column_index_from_string, get_column_letter
    c1 = column_index_from_string(RANGE_COLS[0])
    c2 = column_index_from_string(RANGE_COLS[1])
    max_row = ws.max_row or START_ROW
    addrs = []
    for row in range(START_ROW, max_row + 1):
        for col in range(c1, c2 + 1):
            addrs.append(f"{get_column_letter(col)}{row}")
    return addrs

def _struct_hash(wb: openpyxl.Workbook) -> str:
    """Empreinte globale de structure : noms de feuilles, protections, validations, contenu ligne 1 (questions)."""
    parts = []
    for ws in wb.worksheets:
        parts.append(f"[SHEET]{ws.title}")
        prot = getattr(ws, "protection", None)
        if prot:
            parts.append(f"prot:{bool(prot.sheet)}")
        dvs = getattr(ws, "data_validations", None)
        if dvs and getattr(dvs, "dataValidation", None):
            dv_list = []
            for dv in dvs.dataValidation:
                sqref = str(getattr(dv, "sqref", "")) or ""
                dv_list.append(sqref)
            parts.append("dv:" + "|".join(sorted(dv_list)))
        vals = []
        for cell in ws[1]:
            vals.append(str(cell.value) if cell.value is not None else "")
        parts.append("row1:" + "|".join(vals))
    return _h("\n".join(parts))

def _ensure_sig_sheet(wb: openpyxl.Workbook):
    ws = wb[SIG_SHEET] if SIG_SHEET in wb.sheetnames else wb.create_sheet(SIG_SHEET)
    ws.sheet_state = "veryHidden"
    return ws

# ------------- API -------------
def stamp_workbook(wb: openpyxl.Workbook, *, template_version: str, student_id: str, main_sheet_name: str) -> None:
    """
    Ajoute la feuille _sig et signe chaque cellule utile + header.
    A appeler pendant la génération des copies.
    """
    ws_main = wb[main_sheet_name]
    ws_sig = _ensure_sig_sheet(wb)

    sig_map: Dict[str, str] = {}
    for addr in _cell_range(ws_main):
        v = ws_main[addr].value
        payload = f"{template_version}|{ws_main.title}!{addr}|{type(v).__name__}|{'' if v is None else str(v)}"
        sig_map[addr] = _hmac(payload)

    header = {
        "template_version": template_version,
        "struct_hash": _struct_hash(wb),
        "student_id": student_id,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    ws_sig["A1"] = HEADER_KEY
    ws_sig["B1"] = json.dumps(header)

    # dump des signatures à partir de la ligne 2 : A2=addr, B2=sig
    r = 2
    for addr, sig in sig_map.items():
        ws_sig[f"A{r}"] = addr
        ws_sig[f"B{r}"] = sig
        r += 1

def verify_workbook(path: str, *, main_sheet_name: str) -> Tuple[Dict, List[str], List[str]]:
    """
    Vérifie une copie :
      - retourne (header, cells_changed, issues)
      - cells_changed : liste d'adresses dont la HMAC ne colle plus
      - issues : struct mismatch, _sig manquante, header absent, etc.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    issues: List[str] = []
    if SIG_SHEET not in wb.sheetnames:
        return ({}, [], ["_sig absente (suppression/altération)"])

    ws_sig = wb[SIG_SHEET]
    if (ws_sig["A1"].value or "") != HEADER_KEY:
        issues.append("header _sig invalide/absent")
        header = {}
    else:
        try:
            header = json.loads(ws_sig["B1"].value or "{}")
        except Exception:
            header = {}
            issues.append("header _sig illisible")

    # check struct
    try:
        cur_struct = _struct_hash(wb)
        expected_struct = header.get("struct_hash", "")
        if expected_struct and cur_struct != expected_struct:
            issues.append("structure du workbook modifiée (struct_hash mismatch)")
    except Exception:
        issues.append("erreur calcul struct_hash")

    # reconstruire map signatures
    sig_map = {}
    r = 2
    while True:
        a = ws_sig[f"A{r}"].value
        b = ws_sig[f"B{r}"].value
        if not a:
            break
        sig_map[str(a)] = str(b) if b is not None else ""
        r += 1

    # compare HMAC cellule par cellule
    ws_main = wb[main_sheet_name]
    tv = header.get("template_version", "?")
    changed: List[str] = []
    for addr in sig_map.keys():
        v = ws_main[addr].value
        payload = f"{tv}|{ws_main.title}!{addr}|{type(v).__name__}|{'' if v is None else str(v)}"
        cur = _hmac(payload)
        if cur != sig_map[addr]:
            changed.append(addr)

    return (header, changed, issues)

# app_etudiant.py — version .xlsm (macro-enabled)

import os
import json
import hashlib
import datetime
import openpyxl
import streamlit as st
import re

from auth import (
    get_conn,
    record_submission,
    list_submissions_by_user,
    change_password,
)

DATA_DIR      = os.environ.get("DATA_DIR", "/tmp")  # même valeur que côté prof
GLOBAL_COPIES = os.path.join(DATA_DIR, "copies_generees")
CLASS_ROOT    = os.path.join(DATA_DIR, "classes")
DEPOSIT_DIR   = os.path.join(DATA_DIR, "copies_etudiants")
NOTIF_PATH    = os.path.join(DATA_DIR, "notif_depot.json")

os.makedirs(GLOBAL_COPIES, exist_ok=True)
os.makedirs(DEPOSIT_DIR, exist_ok=True)
if not os.path.exists(NOTIF_PATH):
    with open(NOTIF_PATH, "w", encoding="utf-8") as f:
        json.dump([], f)

STUDENT_CSS = """
<style>
.block-container{ padding: 14px 26px !important; }
html,body,[class*="css"]{ font-family: Inter, system-ui, -apple-system, "Segoe UI", Roboto, sans-serif; }
.student-head .title{ color:#0f172a; font-weight:900; font-size:1.6rem; letter-spacing:.2px; display:flex; align-items:baseline; gap:.5rem; flex-wrap:wrap; margin: 6px 0 4px; }
.student-head .subtitle{ color:#475569; margin:0 0 12px; font-weight:600; }
.student-head .id-badge{ background:linear-gradient(90deg,#7C4DFF,#3FA9F5); color:#fff; font-weight:800; padding:.16rem .55rem; border-radius:10px; }
.student-head .class-badge{ background:linear-gradient(90deg,#10b981,#06b6d4); color:#002a2a; font-weight:900; padding:.16rem .6rem; border-radius:10px; border:1px solid rgba(0,0,0,.06); }
.stTabs [role="tablist"]{ gap:.35rem; border-bottom:1px solid #e5e7eb; padding-bottom:4px; }
.stTabs [role="tab"]{ background:#f8fafc; color:#0f172a; border:1px solid #e5e7eb; border-bottom:none; padding:.5rem .8rem; border-radius:10px 10px 0 0; }
.stTabs [role="tab"][aria-selected="true"]{ background:linear-gradient(90deg,#eef2ff,#e0e7ff); color:#1e293b; font-weight:800; border-color:#c7d2fe; }
</style>
"""

def _slugify(name: str) -> str:
    s = (name or "").lower().strip()
    s = re.sub(r"[^a-z0-9\-_\s]", "", s)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s or "classe"

def _password_errors(pwd: str) -> list[str]:
    """Règles: ≥8, au moins 1 majuscule, 1 minuscule."""
    err = []
    if len(pwd or "") < 8: err.append("au moins 8 caractères")
    if not re.search(r"[A-Z]", pwd or ""): err.append("au moins une majuscule")
    if not re.search(r"[a-z]", pwd or ""): err.append("au moins une minuscule")
    return err

def _copy_filename_for(user_id: str, first_name: str, last_name: str) -> str:
    # <<< .xlsm
    def clean(s: str) -> str:
        return "".join(c for c in s if c.isalnum() or c in ("-", "_"))
    return f"{user_id}_{clean(last_name)}_{clean(first_name)}.xlsm"

def _candidate_copy_paths(user: dict) -> list[str]:
    # Cherche la copie dans copies_generees/ et dans classes/<slug>/copies_generees/
    fn_prefix = user["id"]
    paths = []
    # global
    for f in os.listdir(GLOBAL_COPIES):
        if f.startswith(fn_prefix) and f.endswith(".xlsm"):  # <<< .xlsm
            paths.append(os.path.join(GLOBAL_COPIES, f))
    # par classe
    cls = user.get("class_name") or ""
    if cls:
        slug = _slugify(cls)
        cdir = os.path.join(CLASS_ROOT, slug, "copies_generees")
        if os.path.isdir(cdir):
            for f in os.listdir(cdir):
                if f.startswith(fn_prefix) and f.endswith(".xlsm"):  # <<< .xlsm
                    paths.append(os.path.join(cdir, f))
    return paths

def run(user):
    st.markdown(STUDENT_CSS, unsafe_allow_html=True)

    classe = user.get("class_name") or "—"
    st.markdown(
        f"""
        <div class="student-head">
          <div class="title">
            🎓 Espace Étudiant — {user['first_name']} {user['last_name']}
            <span class="id-badge">{user['id']}</span>
            <span class="class-badge">Classe : {classe}</span>
          </div>
          <div class="subtitle">Gérez votre copie, déposez votre fichier et suivez vos envois.</div>
        </div>
        """,
        unsafe_allow_html=True
    )

    try:
        st.sidebar.info(f"👥 Classe : {classe}")
    except Exception:
        pass

    tabs = st.tabs(["📥 Ma copie", "📤 Dépôt", "🕓 Historique", "🔒 Mot de passe"])

    # --- Ma copie ---
    with tabs[0]:
        paths = _candidate_copy_paths(user)
        if paths:
            p = paths[0]
            st.success("Votre copie est prête.")
            with open(p, "rb") as f:
                st.download_button(
                    label=f"📥 Télécharger ma copie ({os.path.basename(p)})",
                    data=f,
                    file_name=os.path.basename(p),
                    # MIME macro-enabled
                    mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                )
        else:
            st.warning("❗ Aucune copie trouvée pour votre ID. Demandez au professeur de la générer.")

    # --- Dépôt ---
    with tabs[1]:
        st.subheader("📤 Déposer votre fichier Excel rempli")
        fichier_upload = st.file_uploader("Déposez votre fichier (.xlsm)", type=["xlsm"])  # <<< .xlsm

        if fichier_upload:
            try:
                if getattr(fichier_upload, "size", 0) and fichier_upload.size > 20 * 1024 * 1024:
                    st.error("Le fichier est trop volumineux (>20 Mo).")
                    return

                # openpyxl gère .xlsm
                wb = openpyxl.load_workbook(fichier_upload, data_only=True, keep_vba=True)
                ws = wb.active
                id_z1 = ws["Z1"].value
                hash_z2 = ws["Z2"].value

                if not id_z1 or str(id_z1).strip() != user["id"]:
                    st.error("❌ Le fichier ne vous appartient pas (Z1 ≠ votre ID).")
                    return

                # Recalcul du hash comme côté prof
                contenu = str(id_z1).encode()
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            contenu += str(cell).encode()
                recalculated_hash = hashlib.sha256(contenu).hexdigest()

                # Nom standard macro-enabled
                nom_standard = _copy_filename_for(user["id"], user["first_name"], user["last_name"])  # <<< .xlsm
                ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                final_name = f"{ts}__{nom_standard}"

                save_path = os.path.join(DEPOSIT_DIR, final_name)
                with open(save_path, "wb") as out_file:
                    out_file.write(fichier_upload.getbuffer())

                # Notif pour l'espace prof
                try:
                    with open(NOTIF_PATH, "r", encoding="utf-8") as f:
                        notifs = json.load(f)
                except Exception:
                    notifs = []
                if final_name not in notifs:
                    notifs.append(final_name)
                    with open(NOTIF_PATH, "w", encoding="utf-8") as f:
                        json.dump(notifs, f)

                # Trace en BD
                conn = get_conn()
                record_submission(conn, user["id"], final_name, status="received")

                st.success("✅ Dépôt effectué avec succès.")
                with st.expander("Détails techniques (optionnel)"):
                    st.write(f"ID (Z1) : {id_z1}")
                    st.write(f"Hash (Z2) : {hash_z2}")
                    st.write(f"Hash recalculé : {recalculated_hash}")
                    st.write(f"Nom enregistré : {final_name}")

            except Exception as e:
                st.error("❌ Erreur lors du traitement de votre fichier.")
                st.exception(e)

    # --- Historique ---
    with tabs[2]:
        conn = get_conn()
        rows = list_submissions_by_user(conn, user["id"])
        if not rows:
            st.info("Aucun dépôt enregistré pour l’instant.")
        else:
            st.write(f"Vous avez **{len(rows)}** dépôt(s) :")
            for r in rows:
                st.write(f"- {r['submitted_at']} — {r['filename']} — {r['status']}")

    # --- Mot de passe ---
    with tabs[3]:
        st.subheader("🔒 Changer mon mot de passe")
        cur = st.text_input("Mot de passe actuel", type="password")
        new1 = st.text_input("Nouveau mot de passe", type="password")
        new2 = st.text_input("Confirmer le nouveau mot de passe", type="password")

        if st.button("Mettre à jour", key="btn_update_pwd_student"):
            if not cur or not new1 or not new2:
                st.error("Veuillez remplir tous les champs.")
            elif new1 != new2:
                st.error("La confirmation ne correspond pas.")
            else:
                errs = _password_errors(new1)
                if errs:
                    st.error("Le mot de passe doit contenir : " + ", ".join(errs) + ".")
                else:
                    ok = change_password(get_conn(), user["id"], cur, new1)
                    if ok:
                        st.success("✅ Mot de passe mis à jour.")
                    else:
                        st.error("Mot de passe actuel incorrect.")

# compare_excels.py — analyse + intégrité (_sig + HMAC) + Copier-coller + IA + Historique + LOG VBA
# -*- coding: utf-8 -*-

import os, re, csv, json, hashlib, unicodedata, difflib
from datetime import datetime
from collections import defaultdict

import openpyxl
import openpyxl.utils
import pandas as pd
import numpy as np

# --- IA (optionnelle) : mode dégradé si sklearn n'est pas dispo
try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    _SK_OK = True
except Exception:
    TfidfVectorizer = None
    cosine_similarity = None
    _SK_OK = False

# --- Intégrité (_sig)
from integrity import verify_workbook  # verify_workbook(path, main_sheet_name) -> (header, changed_cells, issues)

# --- Cloud (helpers facultatifs)
_SUPA_OK = False
try:
    from supa import upload_file, signed_url
    _SUPA_OK = True
except Exception:
    upload_file = None
    signed_url = None
    _SUPA_OK = False

# ======================= CONFIG =======================
DATA_DIR        = os.environ.get("DATA_DIR", "/tmp")  # même valeur que app_prof
TEMPLATE_PATH   = os.path.join(DATA_DIR, "Fichier_Excel_Professeur_Template.xlsm")  # .xlsm
copies_folder   = os.path.join(DATA_DIR, "copies_etudiants")
rapport_folder  = os.path.join(DATA_DIR, "rapports_etudiants")
hash_log_file   = os.path.join(DATA_DIR, "hash_records.csv")
classes_root    = os.path.join(DATA_DIR, "classes")
cours_file      = os.path.join(DATA_DIR, "cours_references.txt")
dataset_ia_file = os.path.join(DATA_DIR, "dataset.csv")
modifs_csv      = os.path.join(DATA_DIR, "modifications_log_secure.csv")
history_folder  = os.path.join(DATA_DIR, "historique_reponses")
os.makedirs(rapport_folder, exist_ok=True)
os.makedirs(history_folder, exist_ok=True)

# ----- Seuils/réglages
AI_THRESHOLD_DEFAULT   = float(os.environ.get("AI_THRESHOLD_DEFAULT", 0.75))
AI_THRESHOLD_LOWERED   = float(os.environ.get("AI_THRESHOLD_LOWERED", 0.65))
COURSE_RATIO_THRESHOLD = float(os.environ.get("COURSE_RATIO_THRESHOLD", 0.88))
COURSE_LONGEST_MIN     = int(os.environ.get("COURSE_LONGEST_MIN", 70))
FAST_PASTE_SECS        = int(os.environ.get("FAST_PASTE_SECS", 120))
PASTE_MIN_LEN          = int(os.environ.get("PASTE_MIN_LEN", 80))
AI_SHOW_ALL_TABLE      = True

# ======================= HASH INDEX =======================
def _parse_hash_log(path):
    rows = []
    if not os.path.exists(path):
        return rows
    try:
        with open(path, newline="", encoding="utf-8") as f:
            r = csv.DictReader(f)
            for row in r:
                rid = (row.get("id_etudiant") or "").strip()
                h   = (row.get("hash") or "").strip()
                if rid and h:
                    rows.append({
                        "id": rid,
                        "nom": (row.get("nom") or "").strip(),
                        "prenom": (row.get("prenom") or "").strip(),
                        "hash": h,
                        "nom_fichier": (row.get("nom_fichier") or "").strip()
                    })
    except Exception as e:
        print(f"[WARN] Lecture hash log '{path}' impossible : {e}")
    return rows

def _load_all_hash_logs():
    merged = []
    merged.extend(_parse_hash_log(hash_log_file))
    if os.path.isdir(classes_root):
        for slug in os.listdir(classes_root):
            d = os.path.join(classes_root, slug)
            if not os.path.isdir(d):
                continue
            for fname in os.listdir(d):
                if fname.startswith("hash_records_") and fname.endswith(".csv"):
                    merged.extend(_parse_hash_log(os.path.join(d, fname)))
    return merged

def _official_hashes_by_id():
    rows = _load_all_hash_logs()
    m = defaultdict(set)
    for r in rows:
        m[r["id"]].add(r["hash"])
    return m

# ======================= COURS & DATASET IA =======================
cours_content = ""
if os.path.exists(cours_file):
    try:
        with open(cours_file, "r", encoding="utf-8") as f:
            cours_content = f.read().lower()
    except Exception:
        pass

df_ia = None
vectorizer = None
tfidf_matrix = None
if _SK_OK and os.path.exists(dataset_ia_file):
    try:
        df_ia = pd.read_csv(dataset_ia_file)
        if "reponse" in df_ia.columns:
            vectorizer = TfidfVectorizer(stop_words="french")
            tfidf_matrix = vectorizer.fit_transform(df_ia["reponse"].astype(str).fillna(""))
    except Exception as e:
        print(f"[WARN] Erreur chargement dataset IA: {e}")
        df_ia, vectorizer, tfidf_matrix = None, None, None

# ======================= UTILITAIRES =======================
def _safe_str(v):
    try:
        if v is None:
            return ""
        return str(v)
    except Exception:
        return ""

def _excerpt(s: str, n=220) -> str:
    s = _safe_str(s)
    return s if len(s) <= n else (s[:n-1].rstrip() + "…")

def _html_escape(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

def _norm(s: str) -> str:
    if not s: return ""
    s = s.lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def _human_delta(prev_iso: str, now_dt: datetime) -> str:
    if not prev_iso: return ""
    try:
        prev = datetime.fromisoformat(prev_iso)
        return str(now_dt - prev)
    except Exception:
        return ""

def _seconds_since(prev_iso: str, now_dt: datetime) -> float | None:
    if not prev_iso: return None
    try:
        prev = datetime.fromisoformat(prev_iso)
        return (now_dt - prev).total_seconds()
    except Exception:
        return None

def _parse_expected_id_from_filename(nom_fichier: str) -> str | None:
    if "__" in nom_fichier:
        try:
            after = nom_fichier.split("__", 1)[1]
            return after.split("_", 1)[0]
        except Exception:
            return None
    m = re.search(r"(ETUD\d{3,})", nom_fichier.upper())
    return m.group(1) if m else None

def recalculer_hash_depuis_contenu(ws, id_etudiant):
    contenu = (id_etudiant or "").encode()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                contenu += _safe_str(cell).encode()
    return hashlib.sha256(contenu).hexdigest()

# -------- Historique JSON --------
def _history_path(student_id: str) -> str:
    sid = (student_id or "").strip() or "unknown"
    return os.path.join(history_folder, f"{sid}.json")

def _load_history(student_id: str):
    p = _history_path(student_id)
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def _save_history(student_id: str, history_list: list):
    p = _history_path(student_id)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(history_list, f, ensure_ascii=False, indent=2)

def _snapshot_ws(ws, include_cols=(3, 25)) -> dict:
    out = {}
    max_row = ws.max_row or 2
    min_col, max_col = include_cols
    for row in range(2, max_row + 1):
        for col in range(min_col, max_col + 1):
            addr = f"{openpyxl.utils.get_column_letter(col)}{row}"
            out[addr] = _safe_str(ws.cell(row=row, column=col).value)
    return out

# -------- Journal CSV --------
_CSV_HEADER = [
    "timestamp", "time_since_last", "fichier", "id_etudiant",
    "cellule", "question",
    "valeur_avant", "valeur_prof", "valeur_etudiant",
    "source_diff", "action_type", "detection",
    "hash_z2", "hash_recalcule", "tentative_index"
]

def _append_modif_csv(row_dict: dict):
    file_exists = os.path.exists(modifs_csv)
    try:
        with open(modifs_csv, "a", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=_CSV_HEADER)
            if not file_exists:
                w.writeheader()
            w.writerow({k: row_dict.get(k, "") for k in _CSV_HEADER})
    except Exception as e:
        print("[WARN] Erreur écriture journal modifs:", e)

# ======================= DÉTECTION =======================
_AI_MARKERS = [
    "en tant que","dans le cadre","il est important de noter","cependant","par conséquent","de plus",
    "d'après","selon","il convient de","dans un premier temps","notamment","globalement","en conclusion",
    "dans ce contexte","par ailleurs","en outre","autrement dit","de manière générale"
]
_SMART_CHARS = {"×":"signe de multiplication","•":"puce","–":"tiret demi-cadratin","—":"tiret cadratin",
                "’":"apostrophe courbe","“":"guillemet ouvrant","”":"guillemet fermant","\u00A0":"espace insécable",
                "\u202F":"espace fine insécable","…":"points de suspension typographiques"}

def _smart_punct_info(text: str):
    if not text: return (False, "")
    found = [name for ch,name in _SMART_CHARS.items() if ch in text]
    return (len(found) > 0, ", ".join(found))

def _copy_paste_scores(rlow: str):
    if not cours_content: return (0.0, 0)
    seq = difflib.SequenceMatcher(None, rlow, cours_content)
    ratio = float(seq.ratio())
    longest = seq.find_longest_match(0, len(rlow), 0, len(cours_content)).size
    return (ratio, int(longest))

def _looks_paste_burst(text: str) -> bool:
    if not text: return False
    t = text.strip()
    if len(t) < PASTE_MIN_LEN: return False
    return ("\n" in t) or ("  " in t) or ("’" in t) or ("–" in t) or ("•" in t)

def _ai_probability(text: str) -> float:
    if not text: return 0.0
    if _SK_OK and df_ia is not None and vectorizer is not None and tfidf_matrix is not None:
        try:
            vec = vectorizer.transform([text])
            sims = cosine_similarity(vec, tfidf_matrix)
            return float(np.max(sims))
        except Exception:
            pass
    t = _norm(text)
    k = sum(1 for kw in _AI_MARKERS if kw in t)
    nb_sent = t.count(".") + t.count(";") + t.count("!")
    nb_commas = t.count(",")
    long_txt = len(t) >= 140
    score = 0.18*min(k,4) + (0.22 if long_txt else 0) + (0.12 if nb_sent>=2 else 0) + (0.10 if nb_commas>=3 else 0)
    score += 0.08 if "par exemple" in t else 0
    return max(0.0, min(1.0, score))

def _classify(reponse: str, question: str, delta_secs: float | None, prev_text: str | None) -> dict:
    res = {"empty":False,"copy":False,"copy_pct":0,"copy_reason":"","ai":False,"ai_pct":0,"ai_reason":"",
           "ai_score":0,"label":""}
    txt = _safe_str(reponse).strip()
    if txt == "":
        res["empty"] = True; res["label"] = "Non répondu"; return res
    rlow = txt.lower()

    ratio, longest = _copy_paste_scores(rlow)
    paste_like = _looks_paste_burst(txt)
    smart_punct, smart_names = _smart_punct_info(txt)
    fast_paste = (delta_secs is not None and len(txt) >= 60 and delta_secs <= FAST_PASTE_SECS)

    copy_flags = []
    if ratio >= COURSE_RATIO_THRESHOLD or longest >= COURSE_LONGEST_MIN:
        copy_flags.append(f"Très proche du cours (segment commun ≈ {longest} car.)")
    if paste_like: copy_flags.append("Format/longueur typiques d’un collage")
    if smart_punct: copy_flags.append(f"Typographie collée ({smart_names})")
    if fast_paste: copy_flags.append(f"Ajout rapide (< {FAST_PASTE_SECS//60} min)")

    if copy_flags:
        res["copy"] = True
        base = int(round(max(ratio, min(0.99, longest / max(1, len(_norm(rlow))))) * 100))
        bonus = (15 if paste_like else 0) + (12 if smart_punct else 0) + (20 if fast_paste else 0)
        res["copy_pct"] = min(max(base, 55) + bonus, 99)
        res["copy_reason"] = " ; ".join(copy_flags)

    p_ai = _ai_probability(txt)
    res["ai_score"] = int(round(p_ai * 100))
    lowered = (len(_norm(txt)) >= 120) or any(kw in _norm(txt) for kw in _AI_MARKERS)
    ai_threshold = AI_THRESHOLD_LOWERED if lowered else AI_THRESHOLD_DEFAULT
    if p_ai >= ai_threshold:
        res["ai"] = True; res["ai_pct"] = res["ai_score"]; res["ai_reason"] = "Style/longueur/similarité caractéristiques"

    if res["ai"]: res["label"] = f"IA probable ({res['ai_pct']}%)"
    elif res["copy"]: res["label"] = f"Copier-coller (~{res['copy_pct']}%)"
    else: res["label"] = "Réponse normale"
    return res

# ======================= LOG VBA =======================
def _read_embedded_vba_log(wb) -> list[dict]:
    def _s(v): return "" if v is None else str(v)
    def parse_one(ws):
        headers = [(_s(c.value).strip().lower()) for c in ws[1]]
        idx = {h: i + 1 for i, h in enumerate(headers) if h}
        def col(row, *names):
            for key in names:
                j = idx.get(key)
                if j: return _s(ws.cell(row=row, column=j).value)
            return ""
        out = []
        max_row = ws.max_row or 1
        for r in range(2, max_row + 1):
            if not any(_s(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)):
                continue
            ts   = col(r,"horodatage","timestamp","time","date")
            cell = col(r,"cellule","cell")
            q    = col(r,"question")
            oldv = col(r,"avant","old_value","old","valeur_avant")
            newv = col(r,"après","apres","new_value","new","valeur_apres","valeur_etudiant")
            act  = col(r,"action")
            wp   = col(r,"collage","waspaste")
            if not (ts or cell or oldv or newv or act): continue
            if oldv == "" and newv != "": action = "ajout"
            elif oldv != "" and newv == "": action = "suppression"
            elif oldv != newv: action = "modification"
            else: action = act or "inchangé"
            was_paste = (_s(wp).strip().lower() in ("true","vrai","1","oui","yes"))
            out.append({
                "timestamp": ts, "cell": cell, "question": q,
                "old_value": oldv, "new_value": newv,
                "wasPaste": was_paste, "selCount": col(r,"selcount"), "action": action,
            })
        return out

    all_logs = []
    for name in wb.sheetnames:
        if "log" in name.lower():
            try:
                all_logs.extend(parse_one(wb[name]))
            except Exception:
                continue
    return all_logs

# ======================= COEUR =======================
def comparer_etudiant(fichier_etudiant: str) -> str:
    nom_fichier = os.path.basename(fichier_etudiant)
    official_by_id = _official_hashes_by_id()
    expected_id = _parse_expected_id_from_filename(nom_fichier)

    # Ouverture (.xlsm)
    try:
        wb_prof = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True, keep_vba=True)
        wb_etud = openpyxl.load_workbook(fichier_etudiant, data_only=True, keep_vba=True)
        ws_prof = wb_prof.active; ws_etud = wb_etud.active
    except Exception as e:
        return f"❌ Erreur d'ouverture des fichiers : {e}"

    # Identité & Hash
    id_cell = _safe_str(ws_etud["Z1"].value)
    hash_cell = _safe_str(ws_etud["Z2"].value)
    hash_calcule = recalculer_hash_depuis_contenu(ws_etud, id_cell)
    now_dt = datetime.now(); now = now_dt.strftime("%Y-%m-%d %H:%M:%S")

    # Authenticité
    authenticity = "unknown"; authenticity_msg = ""
    official_ok_for_id = id_cell and (hash_cell in official_by_id.get(id_cell, set()))
    if not id_cell or not hash_cell:
        authenticity, authenticity_msg = "critical", "❌ L'ID (Z1) ou le hash (Z2) est manquant."
    else:
        if expected_id and expected_id != id_cell:
            authenticity, authenticity_msg = "mismatch", f"⚠️ ID du fichier (Z1={id_cell}) ≠ ID attendu ({expected_id})."
        if official_ok_for_id and hash_calcule == hash_cell:
            authenticity, authenticity_msg = "official_clean", "✅ Copie officielle intacte."
        elif official_ok_for_id and hash_calcule != hash_cell:
            authenticity, authenticity_msg = "official_then_edited", "✅ Copie officielle utilisée puis contenu modifié (normal)."
        elif (not official_ok_for_id) and hash_calcule == hash_cell:
            authenticity, authenticity_msg = "self_consistent_non_official", "⚠️ Fichier cohérent mais non reconnu parmi les copies officielles."
        elif (not official_ok_for_id) and hash_calcule != hash_cell:
            authenticity, authenticity_msg = "tampered", "🚨 Incohérence : Z2 ≠ contenu et Z2 non-officiel."

    # Colonnes actives (C..Y)
    questions, active_cols = {}, []
    for col in range(3, 26):
        addr = f"{openpyxl.utils.get_column_letter(col)}1"
        qtext = _safe_str(ws_prof[addr].value).strip()
        if qtext:
            questions[addr[:-1]] = qtext; active_cols.append(col)

    # Intégrité (_sig)
    header_sig, changed_cells_sig, issues_sig = verify_workbook(fichier_etudiant, main_sheet_name=ws_etud.title)
    if issues_sig:
        authenticity_msg += (" | " if authenticity_msg else "") + "⚠ Intégrité: " + "; ".join(issues_sig)

    # Journal INTEGRITY
    def _append_integrity(addr):
        m = re.match(r"([A-Z]+)", addr or "")
        col_letter = m.group(1) if m else ""
        question = questions.get(col_letter, "")
        _append_modif_csv({
            "timestamp": now, "time_since_last": "", "fichier": nom_fichier, "id_etudiant": id_cell,
            "cellule": addr, "question": question, "valeur_avant": "", "valeur_prof": "",
            "valeur_etudiant": _safe_str(ws_etud[addr].value) if addr else "",
            "source_diff": "INTEGRITY", "action_type": "modification",
            "detection": "Modifiée après génération (signature)",
            "hash_z2": hash_cell, "hash_recalcule": hash_calcule, "tentative_index": "",
        })
    for addr in changed_cells_sig: _append_integrity(addr)

    # Historique / diffs vs précédent
    hist_key = id_cell or expected_id or "unknown"
    history_list = _load_history(hist_key)
    attempt_index = len(history_list) + 1
    last_ts = history_list[-1]["timestamp"] if history_list else None
    delta_since_last = _human_delta(last_ts, now_dt)
    delta_secs = _seconds_since(last_ts, now_dt)

    snapshot_values = _snapshot_ws(ws_etud)
    prev_values = history_list[-1]["values"] if history_list else {}

    diffs_vs_prev = []
    matrix_full, rows_copy, rows_ai, rows_ai_all = [], [], [], []
    unanswered_count = answered_count = 0

    diffs_vs_template = []
    from collections import defaultdict as _dd
    timeline = _dd(list)

    max_row = max(ws_prof.max_row, ws_etud.max_row)
    for row in range(2, max_row + 1):
        for col in active_cols:
            addr = f"{openpyxl.utils.get_column_letter(col)}{row}"
            col_letter = openpyxl.utils.get_column_letter(col)
            question = questions.get(col_letter, "")
            v_prof = _safe_str(ws_prof.cell(row=row, column=col).value)
            v_etud = _safe_str(ws_etud.cell(row=row, column=col).value)
            prev_text = _safe_str(prev_values.get(addr, "")) if prev_values else ""

            analysis = _classify(v_etud, question, delta_secs, prev_text)
            timeline[addr].append((now, v_etud))

            if analysis["empty"]:
                label = "Non répondu"; unanswered_count += 1
            else:
                label = analysis["label"]; answered_count += 1
                if analysis["copy"]:
                    rows_copy.append((addr, question, _excerpt(v_etud),
                                      f"Copier-coller (~{analysis['copy_pct']}%) — {analysis['copy_reason']}"))
                if analysis["ai"]:
                    rows_ai.append((addr, question, _excerpt(v_etud),
                                    f"IA probable ({analysis['ai_pct']}%) — {analysis['ai_reason']}"))
                rows_ai_all.append((addr, question, _excerpt(v_etud), f"{analysis['ai_score']}%"))

            matrix_full.append((addr, question, v_etud, label))

            if v_prof != v_etud:
                diffs_vs_template.append((addr, v_prof, v_etud))
                _append_modif_csv({
                    "timestamp": now, "time_since_last": delta_since_last, "fichier": nom_fichier, "id_etudiant": id_cell,
                    "cellule": addr, "question": question, "valeur_avant": "", "valeur_prof": v_prof,
                    "valeur_etudiant": v_etud, "source_diff": "TEMPLATE", "action_type": "modification",
                    "detection": label, "hash_z2": hash_cell, "hash_recalcule": hash_calcule, "tentative_index": attempt_index,
                })

            if prev_values and v_etud != prev_text:
                if prev_text == "" and v_etud != "": action = "ajout"
                elif prev_text != "" and v_etud == "": action = "suppression"
                else: action = "modification"
                diffs_vs_prev.append((addr, prev_text, v_etud, action))
                _append_modif_csv({
                    "timestamp": now, "time_since_last": delta_since_last, "fichier": nom_fichier, "id_etudiant": id_cell,
                    "cellule": addr, "question": question, "valeur_avant": prev_text, "valeur_prof": "",
                    "valeur_etudiant": v_etud, "source_diff": "PREVIOUS", "action_type": action,
                    "detection": label, "hash_z2": hash_cell, "hash_recalcule": hash_calcule, "tentative_index": attempt_index,
                })

    # Historique (snapshot)
    history_entry = {
        "timestamp": now, "time_since_last": delta_since_last, "filename": nom_fichier,
        "hash_z2": hash_cell, "hash_recalcule": hash_calcule, "authenticity": authenticity, "values": snapshot_values,
    }
    history_list.append(history_entry); _save_history(hist_key, history_list)

    # Timeline reconstituée
    timeline = defaultdict(list)
    for h in history_list:
        ts = h.get("timestamp"); vals = h.get("values", {})
        for cell, v in sorted(vals.items()):
            if not timeline[cell] or timeline[cell][-1][1] != v:
                timeline[cell].append((ts, v))

    # LOG embarqué (VBA)
    embedded_logs = _read_embedded_vba_log(wb_etud)
    for log in embedded_logs:
        _append_modif_csv({
            "timestamp": log["timestamp"] or now, "time_since_last": delta_since_last, "fichier": nom_fichier,
            "id_etudiant": id_cell, "cellule": log["cell"], "question": log["question"],
            "valeur_avant": log["old_value"], "valeur_prof": "", "valeur_etudiant": log["new_value"],
            "source_diff": "EMBEDDED_LOG", "action_type": log["action"],
            "detection": f"wasPaste={'TRUE' if log['wasPaste'] else 'FALSE'}",
            "hash_z2": hash_cell, "hash_recalcule": hash_calcule, "tentative_index": attempt_index,
        })

    total_changes_template = len(diffs_vs_template)
    total_changes_prev = len(diffs_vs_prev)
    total_integrity_cells = len(changed_cells_sig)
    total_alerts = sum(1 for _, _, _, lab in matrix_full if lab not in ["", "Réponse normale", "Non répondu"])

    # ======================= RAPPORT TXT =======================
    txt_lines = []
    txt_lines.append(f"📄 Rapport : {nom_fichier}")
    txt_lines.append(f"📅 Date : {now}")
    if delta_since_last:
        txt_lines.append(f"⏱️ Temps écoulé depuis la tentative précédente : {delta_since_last}")
    txt_lines.append(f"🧑 ID Étudiant (Z1) : {id_cell}")
    if expected_id: txt_lines.append(f"🧾 ID attendu : {expected_id}")
    txt_lines.append(f"🔐 Hash (Z2) : {hash_cell}")
    txt_lines.append(f"🧪 Hash recalculé : {hash_calcule}")
    txt_lines.append(f"🛡️ Authenticité : {authenticity} — {authenticity_msg}\n")
    txt_lines.append("🧩 Intégrité (signature _sig)")
    if issues_sig:
        for iss in issues_sig: txt_lines.append(f"  • ⚠ {iss}")
    else:
        txt_lines.append("  • ✅ Aucune anomalie structurale détectée.")
    if changed_cells_sig:
        txt_lines.append("  • Cellules modifiées après génération : " + ", ".join(changed_cells_sig))

    base = os.path.splitext(nom_fichier)[0]
    path_txt = os.path.join(rapport_folder, f"{base}_rapport.txt")
    try:
        with open(path_txt, "w", encoding="utf-8") as f: f.write("\n".join(txt_lines))
    except Exception as e:
        return f"❌ Erreur écriture rapport TXT : {e}"

    # ======================= RAPPORT HTML =======================
    def badge(text, kind, title=None):
        colors = {"ok":"#10b981","warn":"#fb923c","err":"#ef4444","info":"#3b82f6","muted":"#64748b"}
        ttl = f' title="{_html_escape(title)}"' if title else ""
        return f'<span class="pill" style="background:{colors.get(kind, "#64748b")}"{ttl}>{_html_escape(text)}</span>'

    def auth_badge_of(state, msg):
        if state == "official_clean": return badge("Copie officielle intacte","ok",msg)
        if state == "official_then_edited": return badge("Copie officielle puis modifiée","info",msg)
        if state == "self_consistent_non_official": return badge("Cohérente non-officielle","warn",msg)
        if state == "tampered": return badge("Incohérence / altération","err",msg)
        if state == "mismatch": return badge("ID ≠ attendu","warn",msg)
        if state == "critical": return badge("Infos manquantes","err",msg)
        return badge("Inconnu","muted",msg)

    def html_rows(rows):
        return "\n".join(f"""
          <tr>
            <td><code>{_html_escape(addr)}</code></td>
            <td>{_html_escape(q)}</td>
            <td>{_html_escape(val)}</td>
            <td>{_html_escape(extra)}</td>
          </tr>""" for addr,q,val,extra in rows)

    def html_rows_diff_prev(rows):
        out=[]
        for addr, old_val, new_val, action in rows:
            col = re.match(r"[A-Z]+", addr).group(0)
            q = questions.get(col, "")
            out.append(f"""
              <tr>
                <td><code>{addr}</code></td>
                <td>{_html_escape(q)}</td>
                <td>{_html_escape(_excerpt(old_val))}</td>
                <td>{_html_escape(_excerpt(new_val))}</td>
                <td>{_html_escape(action)}</td>
              </tr>""")
        return "\n".join(out)

    def html_rows_matrix(matrix):
        def kind_of(label):
            if label.startswith("IA probable"): return "err"
            if label.startswith("Copier-coller"): return "err"
            return "muted"
        return "\n".join(f"""
          <tr>
            <td><code>{_html_escape(addr)}</code></td>
            <td>{_html_escape(q)}</td>
            <td>{_html_escape(v)}</td>
            <td style="text-align:center">{("" if lab in ["","Réponse normale","Non répondu"] else badge(lab, kind_of(lab)))}</td>
          </tr>""" for addr,q,v,lab in matrix)

    def html_rows_timeline(timeline_map):
        out=[]
        for addr in sorted(
            timeline_map.keys(),
            key=lambda a: (re.match(r"[A-Z]+", a).group(0),
                           int(re.search(r"(\d+)$", a).group(1)) if re.search(r"(\d+)$", a) else 0)
        ):
            versions = timeline_map[addr]
            col = re.match(r"[A-Z]+", addr).group(0)
            q = questions.get(col,"")
            items = "".join(f"<div><code>{_html_escape(ts)}</code> → {_html_escape(_excerpt(val,180))}</div>" for ts,val in versions)
            out.append(f"<tr><td><code>{_html_escape(addr)}</code></td><td>{_html_escape(q)}</td><td>{items or '<i class=muted>—</i>'}</td></tr>")
        return "\n".join(out)

    def html_rows_embedded(logs):
        return "\n".join(f"""
          <tr>
            <td><code>{_html_escape(l['timestamp'] or '')}</code></td>
            <td><code>{_html_escape(l['cell'])}</code></td>
            <td>{_html_escape(l['question'])}</td>
            <td>{_html_escape(_excerpt(l['old_value']))}</td>
            <td>{_html_escape(_excerpt(l['new_value']))}</td>
            <td>{_html_escape(l['action'])}</td>
            <td>{"Oui" if l["wasPaste"] else "Non"}</td>
          </tr>""" for l in logs)

    issues_html = "" if not issues_sig else "<ul>" + "".join(f"<li>⚠ {_html_escape(x)}</li>" for x in issues_sig) + "</ul>"

    integrity_table = ""
    if changed_cells_sig:
        rows = []
        for addr in changed_cells_sig:
            m = re.match(r"([A-Z]+)", addr or "")
            col = m.group(1) if m else ""
            q = questions.get(col, "")
            curv = _safe_str(ws_etud[addr].value) if addr else ""
            rows.append((addr, q, _excerpt(curv), "Modifiée après génération (signature)"))
        integrity_table = f"""
        <div class="card scroll-x">
          <h2>🧩 Intégrité — Cellules modifiées après génération</h2>
          <table>
            <thead><tr><th>Cellule</th><th>Question</th><th>Valeur actuelle</th><th>Statut</th></tr></thead>
            <tbody>{html_rows(rows)}</tbody>
          </table>
          <div class="muted table-note">Remarque : c'est normal que l'étudiant remplisse ces cellules. Cette section indique simplement qu'elles ont été modifiées après la génération.</div>
        </div>"""

    copy_table = f"""
    <div class="card scroll-x">
      <h2>📋 Copier-coller (textes recopiés)</h2>
      <table><thead><tr><th>Cellule</th><th>Question</th><th>Extrait</th><th>Indice</th></tr></thead>
        <tbody>{html_rows(rows_copy) if rows_copy else ""}</tbody>
      </table>
      {"<div class='muted table-note'>Aucun cas détecté.</div>" if not rows_copy else ""}
    </div>"""

    ai_table = f"""
    <div class="card scroll-x">
      <h2>🤖 Réponses IA probables</h2>
      <table><thead><tr><th>Cellule</th><th>Question</th><th>Extrait</th><th>Score IA</th></tr></thead>
        <tbody>{html_rows(rows_ai) if rows_ai else ""}</tbody>
      </table>
      {"<div class='muted table-note'>Aucun cas détecté.</div>" if not rows_ai else "<div class='muted small'>Score IA basé sur jeu d’exemples (TF-IDF) ou heuristiques.</div>"}
    </div>"""

    ai_all_table = ""
    if AI_SHOW_ALL_TABLE:
        ai_all_table = f"""
        <div class="card scroll-x">
          <h2>🧠 Score IA (toutes les réponses)</h2>
          <table><thead><tr><th>Cellule</th><th>Question</th><th>Extrait</th><th>Score IA</th></tr></thead>
            <tbody>{html_rows(rows_ai_all)}</tbody>
          </table>
        </div>"""

    timeline_table = f"""
    <div class="card scroll-x">
      <h2>🕒 Historique par cellule (toutes les versions)</h2>
      <table><thead><tr><th>Cellule</th><th>Question</th><th>Versions (horodatage → extrait)</th></tr></thead>
        <tbody>{html_rows_timeline(timeline)}</tbody>
      </table>
      <div class="muted small">Le CSV <code>modifications_log_secure.csv</code> liste aussi ces évènements (source_diff = PREVIOUS).</div>
    </div>"""

    html = f"""<!doctype html>
<html lang="fr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Rapport — {_html_escape(nom_fichier)}</title>
<style>
  :root{{--b:#e5e7eb;--bg:#f8fafc;--muted:#64748b}}
  html,body{{margin:0;padding:0;background:var(--bg);color:#0f172a;font-family:system-ui,-apple-system,Segoe UI,Roboto,Inter,sans-serif}}
  .wrap{{max-width:1200px;margin:0 auto;padding:20px}}
  .card{{background:#fff;border:1px solid var(--b);border-radius:14px;padding:14px 16px;margin:12px 0}}
  h1{{margin:.2rem 0 0;font-size:1.35rem}} h2{{margin:.2rem 0 .6rem;font-size:1.1rem}}
  .muted{{color:var(--muted)}} .pill{{display:inline-flex;align-items:center;padding:.28rem .6rem;border-radius:999px;font-weight:800;color:#fff}}
  table{{border-collapse:collapse;width:100%;table-layout:fixed}}
  th,td{{border:1px solid #e2e8f0;padding:.55rem;vertical-align:top;word-break:break-word;overflow-wrap:anywhere;white-space:pre-wrap}}
  th{{background:#f1f5f9;text-align:left;position:sticky;top:0;z-index:1}} tbody tr:nth-child(odd){{background:#fcfcfd}}
  code{{background:#f1f5f9;padding:.1rem .35rem;border-radius:6px}} .grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:12px}}
  .kpi{{background:linear-gradient(180deg,#ffffff,#fbfdff);border:1px solid var(--b);border-radius:12px;padding:12px}}
  .kpi b{{font-size:1.05rem;font-weight:800;word-break:break-word}} .small{{font-size:.92rem}}
  .table-note{{margin:.4rem 0 0}} .scroll-x{{overflow-x:auto}}
</style></head><body><div class="wrap">

<div class="card">
  <h1>📄 Rapport d'analyse</h1>
  <div class="muted small">{_html_escape(now)}</div>
  {"<div class='muted small'>⏱️ Depuis la tentative précédente : " + _html_escape(delta_since_last) + "</div>" if delta_since_last else ""}
</div>

<div class="grid">
  <div class="kpi"><div class="muted">Étudiant (Z1)</div><b>{_html_escape(id_cell)}</b></div>
  <div class="kpi"><div class="muted">Fichier</div><b>{_html_escape(nom_fichier)}</b></div>
  <div class="kpi"><div class="muted">Authenticité</div><b>{auth_badge_of(authenticity, authenticity_msg)}</b><div class="muted small" style="margin-top:.3rem">{_html_escape(authenticity_msg)}</div></div>
  <div class="kpi"><div class="muted">Hash Z2</div><b><code>{_html_escape(hash_cell)}</code></b></div>
  <div class="kpi"><div class="muted">Hash recalculé</div><b><code>{_html_escape(hash_calcule)}</code></b></div>
  <div class="kpi"><div class="muted">Tentative</div><b>#{attempt_index}</b></div>
</div>

<div class="card"><h2>🧩 Intégrité (signature _sig)</h2>{("<div class='muted'>Aucun problème d'intégrité détecté.</div>" if not issues_sig else issues_html)}</div>

{integrity_table}
{copy_table}
{ai_table}
{ai_all_table if AI_SHOW_ALL_TABLE else ""}

<div class="card scroll-x">
  <h2>🔁 Changements depuis la tentative précédente</h2>
  <table><thead><tr><th>Cellule</th><th>Question</th><th>Avant</th><th>Maintenant</th><th>Action</th></tr></thead>
    <tbody>{html_rows_diff_prev(diffs_vs_prev)}</tbody>
  </table>
  {"<div class='muted table-note'>Aucun</div>" if not diffs_vs_prev else ""}
</div>

<div class="card scroll-x">
  <h2>🧰 Traçabilité locale (VBA) — saisies détaillées</h2>
  <table><thead><tr><th>Horodatage</th><th>Cellule</th><th>Question</th><th>Avant</th><th>Après</th><th>Action</th><th>Collage</th></tr></thead>
    <tbody>{html_rows_embedded(embedded_logs) if embedded_logs else ""}</tbody>
  </table>
  {"<div class='muted table-note'>Aucun log embarqué détecté (feuille LOG absente ou vide).</div>" if not embedded_logs else ""}
</div>

<div class="card scroll-x">
  <h2>🗺️ Grille (réponses de l'étudiant)</h2>
  <table><thead><tr><th>Cellule</th><th>Question</th><th>Réponse</th><th>Signal</th></tr></thead>
    <tbody>{html_rows_matrix(matrix_full)}</tbody>
  </table>
  <div class="muted table-note">Les tableaux ci-dessus donnent une vue complète.</div>
</div>

<div class="card scroll-x">
  <h2>🕒 Historique par cellule (toutes les versions)</h2>
  <table><thead><tr><th>Cellule</th><th>Question</th><th>Versions (horodatage → extrait)</th></tr></thead>
    <tbody>{html_rows_timeline(timeline)}</tbody>
  </table>
</div>

<div class="grid">
  <div class="kpi"><div class="muted">Cells ≠ template</div><b>{total_changes_template}</b></div>
  <div class="kpi"><div class="muted">Changements vs précédent</div><b>{total_changes_prev}</b></div>
  <div class="kpi"><div class="muted">Cellules modifiées (signature)</div><b>{total_integrity_cells}</b></div>
  <div class="kpi"><div class="muted">Réponses renseignées</div><b>{answered_count}</b></div>
  <div class="kpi"><div class="muted">Non répondu</div><b>{unanswered_count}</b></div>
  <div class="kpi"><div class="muted">Alertes (grille)</div><b>{total_alerts}</b></div>
</div>

</div></body></html>
"""
    path_html = os.path.join(rapport_folder, f"{base}_rapport.html")
    try:
        with open(path_html, "w", encoding="utf-8") as f: f.write(html)
    except Exception as e:
        return f"❌ Erreur écriture rapport HTML : {e}"

    # -------- Upload Supabase (facultatif) ----------
    cloud_msg = ""
    if _SUPA_OK:
        try:
            student_key = (id_cell or expected_id or "unknown").strip() or "unknown"
            remote_dir = f"rapports/{student_key}/"
            remote_txt = remote_dir + os.path.basename(path_txt)
            remote_htm = remote_dir + os.path.basename(path_html)
            upload_file(path_txt, remote_txt, content_type="text/plain")
            upload_file(path_html, remote_htm, content_type="text/html")
            url_txt = signed_url(remote_txt, expires_in=7*24*3600)
            url_htm = signed_url(remote_htm, expires_in=7*24*3600)
            cloud_msg = f" | cloud: TXT={url_txt} | HTML={url_htm}"
        except Exception as e:
            cloud_msg = f" | cloud: échec upload ({e})"

    return f"📁 Rapports générés : {path_txt} | {path_html}{cloud_msg}"

# ======================= CLI =======================
if __name__ == "__main__":
    print("🔍 Analyse des copies en cours...\n")
    for fichier in os.listdir(copies_folder):
        if fichier.lower().endswith(".xlsm"):
            chemin = os.path.join(copies_folder, fichier)
            print(comparer_etudiant(chemin))
    print("\n✅ Analyse terminée. Rapports dans :", rapport_folder)

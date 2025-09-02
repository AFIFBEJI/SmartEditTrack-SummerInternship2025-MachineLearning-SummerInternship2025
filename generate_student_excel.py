

import openpyxl
import hashlib
import os
import csv
from openpyxl.styles import Protection

# >>> estampillage d’intégrité
from integrity import stamp_workbook  # stamp_workbook(wb, template_version, student_id, main_sheet_name)

# === CONFIGURATION ===
template_path = "Fichier_Excel_Professeur_Template.xlsm"   # <<< .xlsm
students_csv  = "liste_etudiants.csv"
output_folder = "copies_etudiants"
log_file      = "hash_records.csv"

# Option simple pour version du template (ou lis-la depuis une cellule si tu préfères)
TEMPLATE_VERSION = os.environ.get("TEMPLATE_VERSION", "v1")

# === ASSURER DOSSIER DE SORTIE ===
os.makedirs(output_folder, exist_ok=True)

# === CHARGER LISTE DES ÉTUDIANTS ===
etudiants = []
with open(students_csv, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        etudiants.append({
            "id":     (row.get("id") or "").strip(),
            "nom":    (row.get("nom") or "").strip(),
            "prenom": (row.get("prenom") or "").strip(),
        })

# === PRÉPARER FICHIER LOG HASHS ===
with open(log_file, "w", newline="", encoding="utf-8") as f_log:
    writer = csv.writer(f_log)
    writer.writerow(["id_etudiant", "nom", "prenom", "hash", "nom_fichier"])

    # === POUR CHAQUE ÉTUDIANT ===
    for etu in etudiants:
        id_etudiant = etu["id"]
        nom         = etu["nom"]
        prenom      = etu["prenom"]

        if not id_etudiant:
            print("⚠️ Ligne CSV sans id, ignorée.")
            continue

        # Charger le modèle (conserver les macros)
        wb = openpyxl.load_workbook(template_path, keep_vba=True, data_only=False)
        ws = wb.active

        # Écrire ID & futur hash en Z1 / Z2
        ws["Z1"] = id_etudiant

        # Calculer hash basé sur ID + contenu (unicité + traçabilité)
        contenu_concatene = id_etudiant.encode("utf-8")
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    contenu_concatene += str(cell).encode("utf-8")
        hash_etudiant = hashlib.sha256(contenu_concatene).hexdigest()
        ws["Z2"] = hash_etudiant

        # === Masquer la colonne Z (ID/Hash) ===
        ws.column_dimensions["Z"].hidden = True

        # === Déverrouiller C:Y pour saisie et protéger la feuille ===
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row or 2, min_col=3, max_col=25):
            for cell in row_cells:
                cell.protection = Protection(locked=False)

        # Quelques options de protection « confort »
        ws.protection.sheet = True
        ws.protection.enable()
        ws.protection.selectLockedCells = False
        ws.protection.selectUnlockedCells = True
        # (Facultatif) ws.protection.set_password("motdepasse-prof")

        # === Estampiller le classeur (_sig VeryHidden + HMAC par cellule) ===
        stamp_workbook(
            wb,
            template_version=TEMPLATE_VERSION,
            student_id=id_etudiant,
            main_sheet_name=ws.title,
        )

        # Sauvegarder le fichier personnalisé (.xlsm pour conserver les macros)
        nom_fichier   = f"{id_etudiant}_{nom}_{prenom}.xlsm"   # <<< .xlsm
        chemin_sortie = os.path.join(output_folder, nom_fichier)
        wb.save(chemin_sortie)

        # Journaliser
        writer.writerow([id_etudiant, nom, prenom, hash_etudiant, nom_fichier])
        print(f"✅ Copie générée pour {id_etudiant} : {nom_fichier}")
